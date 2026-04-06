"""Ticketing endpoints."""

from __future__ import annotations

import io
from datetime import datetime, timezone
from typing import Annotated

from fastapi import APIRouter, Depends, HTTPException, Query, Response, status
from pydantic import BaseModel, field_validator

from backend.infra.database.repositories.ticket_repository import TicketRepository
from backend.infra.database.repositories.customer_repository import CustomerRepository
from backend.interfaces.api.dependencies import _get_session

router = APIRouter(prefix="/tickets", tags=["tickets"])

_FINAL_STATUSES = {"resolved", "closed"}

# 2 email template phases
_EMAIL_TEMPLATES = {
    "in_progress": {
        "subject": "{task}",
        "body": (
            "Dengan hormat,\n\n"
            "Kami informasikan bahwa permintaan Anda dengan task:\n"
            "  \"{task}\"\n\n"
            "saat ini sedang dalam proses pengerjaan.\n\n"
            "PIC: {pic}\n"
            "Status: In Progress\n\n"
            "Kami akan segera memberikan update lebih lanjut.\n\n"
            "Terima kasih atas kesabaran Anda.\n\n"
            "Salam,\n{pic}"
        ),
    },
    "selesai": {
        "subject": "Re: {task}",
        "body": (
            "Dengan hormat,\n\n"
            "Kami informasikan bahwa task berikut telah diselesaikan:\n"
            "  \"{task}\"\n\n"
            "PIC: {pic}\n"
            "Status: Selesai\n"
            "{solution_line}"
            "\nMohon konfirmasi apabila ada yang perlu ditindaklanjuti.\n\n"
            "Terima kasih atas kepercayaan Anda.\n\n"
            "Salam,\n{pic}"
        ),
    },
}


class TicketResponse(BaseModel):
    id: str
    ticket_no: str
    customer_id: str | None = None
    task: str
    pic: str
    status: str
    description_solution: str | None = None
    created_at: str
    ended_at: str | None = None


class CreateTicketRequest(BaseModel):
    customer_id: str
    task: str
    pic: str
    status: str = "open"
    description_solution: str | None = None

    @field_validator("status")
    @classmethod
    def valid_status(cls, value: str) -> str:
        if value not in {"open", "in_progress", "resolved", "closed"}:
            raise ValueError("status must be one of: open, in_progress, resolved, closed")
        return value

    @field_validator("customer_id")
    @classmethod
    def valid_customer_id(cls, value: str) -> str:
        cleaned = value.strip()
        if not cleaned:
            raise ValueError("customer_id is required")
        return cleaned


class UpdateTicketRequest(BaseModel):
    customer_id: str | None = None
    task: str | None = None
    pic: str | None = None
    status: str | None = None
    description_solution: str | None = None

    @field_validator("status")
    @classmethod
    def valid_status(cls, value: str | None) -> str | None:
        if value is None:
            return value
        if value not in {"open", "in_progress", "resolved", "closed"}:
            raise ValueError("status must be one of: open, in_progress, resolved, closed")
        return value

    @field_validator("customer_id")
    @classmethod
    def valid_optional_customer_id(cls, value: str | None) -> str | None:
        if value is None:
            return value
        cleaned = value.strip()
        if not cleaned:
            raise ValueError("customer_id must not be blank")
        return cleaned


def _to_response(row) -> TicketResponse:
    return TicketResponse(
        id=row.id,
        ticket_no=row.ticket_no,
        customer_id=row.customer_id,
        task=row.task,
        pic=row.pic,
        status=row.status,
        description_solution=row.description_solution,
        created_at=row.created_at.isoformat(),
        ended_at=row.ended_at.isoformat() if row.ended_at else None,
    )


def _get_repo():
    session = _get_session()
    try:
        yield TicketRepository(session)
    finally:
        session.close()


def _get_customer_repo():
    session = _get_session()
    try:
        yield CustomerRepository(session)
    finally:
        session.close()


@router.get("/export")
def export_tickets(
    month: int | None = Query(default=None, ge=1, le=12),
    year: int | None = Query(default=None),
    repo: Annotated[TicketRepository, Depends(_get_repo)] = None,
    customer_repo: Annotated[CustomerRepository, Depends(_get_customer_repo)] = None,
):
    """Export tickets as Excel — one sheet per customer."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    tickets = repo.list_tickets(month=month, year=year)

    # Build customer lookup
    customers = {c.id: c.display_name for c in customer_repo.list_customers()}

    # Group tickets by customer
    from collections import defaultdict
    by_customer: dict[str, list] = defaultdict(list)
    for t in tickets:
        key = t.customer_id or "__none__"
        by_customer[key].append(t)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    HEADERS = ["Task", "PIC", "Status", "Dibuat", "Selesai", "Deskripsi / Solusi"]
    HEADER_FILL = PatternFill("solid", fgColor="1E40AF")
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    STATUS_COLORS = {
        "open": "BFDBFE",
        "in_progress": "FDE68A",
        "resolved": "BBF7D0",
        "closed": "BBF7D0",
    }

    def fmt_dt(val) -> str:
        if val is None:
            return "-"
        if isinstance(val, str):
            try:
                val = datetime.fromisoformat(val)
            except Exception:
                return val
        return val.strftime("%d %b %Y")

    for customer_id, tkts in by_customer.items():
        sheet_name = customers.get(customer_id, "Tanpa Customer")
        # Excel sheet name max 31 chars, no special chars
        safe_name = sheet_name[:31].replace("/", "-").replace("\\", "-").replace("*", "").replace("?", "").replace("[", "").replace("]", "").replace(":", "")
        ws = wb.create_sheet(title=safe_name)

        # Header row
        for col, h in enumerate(HEADERS, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.row_dimensions[1].height = 20

        for row_idx, t in enumerate(tkts, 2):
            status_color = STATUS_COLORS.get(t.status, "FFFFFF")
            row_data = [
                t.task,
                t.pic,
                t.status.replace("_", " ").title(),
                fmt_dt(t.created_at),
                fmt_dt(t.ended_at),
                t.description_solution or "",
            ]
            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                if col_idx == 3:  # status column
                    cell.fill = PatternFill("solid", fgColor=status_color)

        # Column widths
        col_widths = [48, 16, 14, 14, 14, 52]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        ws.freeze_panes = "A2"

    if not wb.sheetnames:
        ws = wb.create_sheet("Tickets")
        ws.append(["Tidak ada data"])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    month_label = f"_{month:02d}" if month else ""
    year_label = f"_{year}" if year else ""
    filename = f"tickets{year_label}{month_label}.xlsx"

    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("", response_model=list[TicketResponse])
def list_tickets(
    customer_id: str | None = Query(default=None),
    month: int | None = Query(default=None, ge=1, le=12),
    year: int | None = Query(default=None),
    repo: Annotated[TicketRepository, Depends(_get_repo)] = None,
):
    return [_to_response(row) for row in repo.list_tickets(
        customer_id=customer_id,
        month=month,
        year=year,
    )]


@router.post("", response_model=TicketResponse, status_code=status.HTTP_201_CREATED)
def create_ticket(
    body: CreateTicketRequest,
    repo: Annotated[TicketRepository, Depends(_get_repo)],
):
    ended_at = datetime.now(timezone.utc) if body.status in _FINAL_STATUSES else None
    row = repo.create_ticket(
        ticket_no=repo.next_ticket_number(),
        customer_id=body.customer_id,
        task=body.task,
        pic=body.pic,
        status=body.status,
        description_solution=body.description_solution,
        ended_at=ended_at,
    )
    repo.session.commit()
    return _to_response(row)


@router.patch("/{ticket_id}", response_model=TicketResponse)
def update_ticket(
    ticket_id: str,
    body: UpdateTicketRequest,
    repo: Annotated[TicketRepository, Depends(_get_repo)],
):
    existing = repo.get_ticket(ticket_id)
    if existing is None:
        raise HTTPException(status_code=404, detail="Ticket not found")

    updates: dict[str, object] = {}
    if body.customer_id is not None:
        updates["customer_id"] = body.customer_id
    if body.task is not None:
        updates["task"] = body.task
    if body.pic is not None:
        updates["pic"] = body.pic
    if body.description_solution is not None:
        updates["description_solution"] = body.description_solution
    if body.status is not None:
        updates["status"] = body.status
        if body.status in _FINAL_STATUSES:
            updates["ended_at"] = existing.ended_at or datetime.now(timezone.utc)
        else:
            updates["ended_at"] = None

    row = repo.update_ticket(ticket_id, **updates)
    if row is None:
        raise HTTPException(status_code=404, detail="Ticket not found")

    repo.session.commit()
    return _to_response(row)


@router.delete("/{ticket_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_ticket(
    ticket_id: str,
    repo: Annotated[TicketRepository, Depends(_get_repo)],
):
    ticket = repo.get_ticket(ticket_id)
    if ticket is None:
        raise HTTPException(status_code=404, detail="Ticket not found")
    repo.session.delete(ticket)
    repo.session.commit()


@router.get("/{ticket_id}/email-template")
def get_email_template(
    ticket_id: str,
    template_type: str = Query(default="in_progress"),
    repo: Annotated[TicketRepository, Depends(_get_repo)] = None,
    customer_repo: Annotated[CustomerRepository, Depends(_get_customer_repo)] = None,
):
    """Generate email template for a ticket. template_type: in_progress | selesai"""
    ticket = repo.get_ticket(ticket_id)
    if ticket is None:
        raise HTTPException(status_code=404, detail="Ticket not found")

    if template_type not in _EMAIL_TEMPLATES:
        raise HTTPException(
            status_code=400,
            detail=f"template_type must be one of: {', '.join(_EMAIL_TEMPLATES.keys())}",
        )

    customer_name = ""
    if ticket.customer_id:
        customer = customer_repo.get_customer(ticket.customer_id)
        if customer:
            customer_name = customer.display_name

    solution_line = (
        f"\nSolusi:\n{ticket.description_solution}\n"
        if ticket.description_solution
        else ""
    )

    tpl = _EMAIL_TEMPLATES[template_type]
    subject = tpl["subject"].format(task=ticket.task)
    body = tpl["body"].format(
        task=ticket.task,
        pic=ticket.pic,
        solution_line=solution_line,
    )

    return {
        "ticket_id": ticket_id,
        "template_type": template_type,
        "customer_name": customer_name,
        "subject": subject,
        "body": body,
    }
